from openpyxl import load_workbook
from openpyxl.packaging import extended
import re
from ..re_pattern import patterns
from io import BytesIO
import datetime

def sheet_process(file, file_id, user, org):
    wb = load_workbook(file)
    for w in wb.worksheets:
        for r in w.iter_rows():
            for cell in r:
                if cell.value is not None:
                    tmp = str(cell.value)
                    for p in patterns:
                        s = re.search(p, tmp)
                        if s:
                            if(len(s.group())<=8):
                                tmp = cell.value.replace(s.group(),'X'*len(s.group()))
                                cell.value = tmp
                                continue
                            tmp = tmp[:s.start()+3] + 'X'*(len(s.group())-6) + tmp[s.end()-3:]
                            cell.value = tmp
    
    time = datetime.datetime.now() - datetime.timedelta(hours=8)
    wb.properties.creator = "{}@{};{}@{}".format(user.username, org.name, user.uid, org.oid)
    wb.properties.title = file_id
    wb.properties.description = "由文件脱敏系统处理部分关键文字信息。\n@Sensitive Data Recognizer\n@Xiamen_Torch_HiTech_IDZ\n@Chengyi University College, Jimei University"
    wb.properties.created = time
    wb.properties.modified = time
    wb.properties.lastModifiedBy = "Sensitive Data Recognizer@Xiamen_Torch_HiTech_IDZ"
    wb.properties.lastPrinted = time

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output